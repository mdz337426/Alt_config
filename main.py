import io
from pdf2image import convert_from_path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
import app
import base64
import re
import os
import ext_examples

prompt = """
Write single alt text for each image for that caption only. Each alt text should include the necessary information given in the image.

Start the alt text with the type of image. For example, an illustration, a diagram, a flowchart, a radiograph, etc. Write flowchart data in bullet points. 

Do not include or rephrase caption in the alt text. Also, do not include the purpose or use of the image in the alt text.

Spell out symbol and units.

Use nurse instead of woman and use word 'patient' for the woman in hospital bed. Avoid mentioning gender of nurses.

Use space between abbreviations. For example, DNA should be D N A.
Spell out symbols and units.

give the alt text of the given Fig. only which caption is...

give response in plain text only ...
Do not write anything except the response it should be raw text only.
"""

caption_prompt ="""
if this page contains any figures or illustrations present and have caption below of format 'Fig\.\s*\d+\.\d+' 


Extract the captions of each figure
if there is more than one figures present with caption in the page make it double newline separated include 

if page contain no figures simply give response -1, ignore the references

give response in plain text only ...
Do not write anything except the response it should be raw text only.
"""


file_prompt = """
This is file is for your training purpose.
"""


def check_uploaded():
    pass


def pdf_images_to_excel(pdf_path, excel_path, completed_pages, row_number, sample_file):
    sample = ext_examples.extract_alt_text_examples(sample_file,num_examples=5)

    # Convert PDF pages to images
    images = convert_from_path(pdf_path)
    # Create a new Excel workbook and sheet
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()

    ws = wb.active
    ws.title = "PDF Images"
    ws[f'A1'] = "Page No."
    ws[f'C1'] = "Figure"
    ws[f'E1'] = "Caption"
    ws[f'G1'] = "ALT-Text"

    try:
        i=row_number
        for j, image in enumerate(images):
            if j<completed_pages:
                continue 
            # Convert image to bytes
            img_buffer = io.BytesIO()
            image.save(img_buffer, format="PNG")
            img_buffer.seek(0)

            # Convert to Base64
            img_base64 = base64.b64encode(img_buffer.read()).decode("utf-8")
            caption = app.get_response(img_base64,caption_prompt)
            if len(caption) <=2 : 
                print(f"page: {j} skipped because no diagrams found")
                continue
            print(caption)
            caption = caption.split("\n\n")
    
            for item in caption: 
                fig = re.match(r'(Fig\.\s*\d+\.\d+)', item)
                if fig:
                    fig = fig.group(0)
                else:
                    fig = "N/A"

                alt_text = app.generate_alt_text(sample, img_base64, prompt+item)                         
                ws[f'A{i+2}'] = f"Page no. {j}"
                ws[f'C{i+2}'] = fig
                text = item[len(fig):].strip()
                ws[f'E{i+2}'] = text+'\n'
                ws[f'G{i+2}'] = alt_text
                i+=1
            print(f"page {j} completed")
        wb.save(excel_path)
        print(f"Excel saved at: {excel_path}")       
    except:
        wb.save(excel_path)
        print(f"Excel saved at: {excel_path}")


     
def main():
    completed_pages = int(input("Enter the number of completed pages: "))
    sapmle_path = "sample.xlsx"
    row_number = int(input("Enter the number of rows completed: "))
    pdf_images_to_excel("input.pdf", "output.xlsx", completed_pages, row_number, sapmle_path)


if __name__ == "__main__":
    main()